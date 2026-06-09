"""Tests for the XLSForm -> Supform schema importer.

Each test builds a minimal in-memory .xlsx workbook with openpyxl, feeds the raw bytes to
import_xlsform(), and asserts on the resulting FormSchema.  No filesystem I/O required.
"""

from __future__ import annotations

import io
from typing import Any

import pytest
from openpyxl import Workbook

from app.importers import import_xlsform
from app.schemas.form_schema import FormSchema

# ------------------------------------------------------------------ helpers


def _xlsx(
    survey: list[list[Any]],
    choices: list[list[Any]] | None = None,
    settings: list[list[Any]] | None = None,
) -> bytes:
    """Build a minimal XLSForm workbook and return its raw bytes."""
    wb = Workbook()
    # openpyxl creates a default 'Sheet' — remove it.
    for name in list(wb.sheetnames):
        del wb[name]

    ws = wb.create_sheet("survey")
    for row in survey:
        ws.append(row)

    if choices:
        wc = wb.create_sheet("choices")
        for row in choices:
            wc.append(row)

    if settings:
        ws2 = wb.create_sheet("settings")
        for row in settings:
            ws2.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _result(
    survey: list[list[Any]],
    choices: list[list[Any]] | None = None,
    settings: list[list[Any]] | None = None,
) -> FormSchema:
    return import_xlsform(_xlsx(survey, choices, settings))


# ------------------------------------------------------------------ basic type mapping


def test_text_and_integer_fields() -> None:
    data = _result(
        survey=[
            ["type", "name", "label"],
            ["text", "full_name", "Full name"],
            ["integer", "age", "Age"],
        ]
    )
    els = data.pages[0].elements
    assert len(els) == 2
    assert els[0].type == "text" and els[0].name == "full_name"
    assert els[1].type == "integer" and els[1].name == "age"


def test_decimal_note_date_time_datetime() -> None:
    data = _result(
        survey=[
            ["type", "name", "label"],
            ["decimal", "score", "Score"],
            ["note", "intro", "Welcome"],
            ["date", "dob", "DOB"],
            ["time", "t", "Time"],
            ["datetime", "ts", "Timestamp"],
        ]
    )
    types = {el.name: el.type for el in data.pages[0].elements}
    assert types == {
        "score": "decimal",
        "intro": "note",
        "dob": "date",
        "t": "time",
        "ts": "datetime",
    }


def test_media_and_calculate_types() -> None:
    data = _result(
        survey=[
            ["type", "name", "label", "calculation"],
            ["image", "photo", "Photo", ""],
            ["audio", "clip", "Clip", ""],
            ["video", "vid", "Video", ""],
            ["calculate", "total", "", "2 + 2"],
        ]
    )
    types = {el.name: el.type for el in data.pages[0].elements}
    assert types["photo"] == "image"
    assert types["clip"] == "file"
    assert types["vid"] == "file"
    assert types["total"] == "calculated"


def test_range_maps_to_scale() -> None:
    data = _result(
        survey=[
            ["type", "name", "label"],
            ["range", "nps", "NPS"],
        ]
    )
    assert data.pages[0].elements[0].type == "scale"


def test_unknown_metadata_types_dropped() -> None:
    data = _result(
        survey=[
            ["type", "name", "label"],
            ["start", "start", ""],
            ["end", "end", ""],
            ["text", "q1", "Q1"],
        ]
    )
    assert len(data.pages[0].elements) == 1
    assert data.pages[0].elements[0].name == "q1"


# ------------------------------------------------------------------ select types


def test_select_one_becomes_single_choice() -> None:
    data = _result(
        survey=[
            ["type", "name", "label"],
            ["select_one colors", "fav_color", "Favourite color"],
        ],
        choices=[
            ["list_name", "name", "label"],
            ["colors", "red", "Red"],
            ["colors", "blue", "Blue"],
        ],
    )
    el = data.pages[0].elements[0]
    assert el.type == "single_choice"
    values = [o.value for o in (el.options or [])]
    assert values == ["red", "blue"]


def test_select_multiple_becomes_multi_choice() -> None:
    data = _result(
        survey=[
            ["type", "name", "label"],
            ["select_multiple langs", "langs", "Languages"],
        ],
        choices=[
            ["list_name", "name", "label"],
            ["langs", "py", "Python"],
            ["langs", "go", "Go"],
        ],
    )
    el = data.pages[0].elements[0]
    assert el.type == "multi_choice"
    assert len(el.options or []) == 2


def test_select_one_space_variant() -> None:
    """'select one list' (space, not underscore) should also work."""
    data = _result(
        survey=[
            ["type", "name", "label"],
            ["select one yesno", "yn", "Yes/No"],
        ],
        choices=[
            ["list_name", "name", "label"],
            ["yesno", "yes", "Yes"],
            ["yesno", "no", "No"],
        ],
    )
    el = data.pages[0].elements[0]
    assert el.type == "single_choice"


# ------------------------------------------------------------------ groups and repeats


def test_begin_end_group_creates_group_container() -> None:
    data = _result(
        survey=[
            ["type", "name", "label"],
            ["begin group", "personal", "Personal info"],
            ["text", "first_name", "First name"],
            ["text", "last_name", "Last name"],
            ["end group", "", ""],
            ["text", "outside", "Outside group"],
        ]
    )
    els = data.pages[0].elements
    assert els[0].type == "group"
    assert els[0].name == "personal"
    assert len(els[0].elements or []) == 2
    assert els[1].name == "outside"


def test_begin_end_repeat_creates_repeat() -> None:
    data = _result(
        survey=[
            ["type", "name", "label"],
            ["begin repeat", "members", "Members"],
            ["text", "member_name", "Name"],
            ["end repeat", "", ""],
        ]
    )
    el = data.pages[0].elements[0]
    assert el.type == "repeat"
    assert el.repeat is not None
    assert len(el.elements or []) == 1


def test_nested_group_inside_group() -> None:
    data = _result(
        survey=[
            ["type", "name", "label"],
            ["begin group", "outer", "Outer"],
            ["begin group", "inner", "Inner"],
            ["text", "q", "Q"],
            ["end group", "", ""],
            ["end group", "", ""],
        ]
    )
    outer = data.pages[0].elements[0]
    assert outer.type == "group"
    inner = (outer.elements or [])[0]
    assert inner.type == "group"
    assert len(inner.elements or []) == 1


# ------------------------------------------------------------------ logic / required / visibleIf


def test_required_column_sets_required_flag() -> None:
    data = _result(
        survey=[
            ["type", "name", "label", "required"],
            ["text", "q", "Q", "yes"],
        ]
    )
    assert data.pages[0].elements[0].required is True


def test_relevant_becomes_visible_if() -> None:
    data = _result(
        survey=[
            ["type", "name", "label", "relevant"],
            ["text", "detail", "Detail", "${age} >= 18"],
        ]
    )
    el = data.pages[0].elements[0]
    assert el.visible_if == "age >= 18"


def test_constraint_becomes_validation_expression() -> None:
    data = _result(
        survey=[
            ["type", "name", "label", "constraint", "constraint_message"],
            ["integer", "score", "Score", ". >= 0 and . <= 100", "0-100 only"],
        ]
    )
    el = data.pages[0].elements[0]
    assert el.validation is not None
    assert "value" in (el.validation.expression or "")
    assert el.validation.message == "0-100 only"


def test_calculate_field_has_calculate_and_readonly() -> None:
    data = _result(
        survey=[
            ["type", "name", "label", "calculation"],
            ["calculate", "total", "", "${a} + ${b}"],
        ]
    )
    el = data.pages[0].elements[0]
    assert el.type == "calculated"
    assert el.calculate == "a + b"
    assert el.read_only is True


# ------------------------------------------------------------------ i18n labels


def test_multilingual_labels_parsed_as_dict() -> None:
    data = _result(
        survey=[
            ["type", "name", "label::English (en)", "label::French (fr)"],
            ["text", "q", "Question", "Question FR"],
        ]
    )
    el = data.pages[0].elements[0]
    assert isinstance(el.label, dict)
    assert el.label.get("en") == "Question"
    assert el.label.get("fr") == "Question FR"


def test_plain_label_string() -> None:
    data = _result(
        survey=[
            ["type", "name", "label"],
            ["text", "q", "Simple label"],
        ]
    )
    el = data.pages[0].elements[0]
    assert el.label == "Simple label"


# ------------------------------------------------------------------ settings


def test_form_title_from_settings() -> None:
    data = _result(
        survey=[
            ["type", "name", "label"],
            ["text", "q", "Q"],
        ],
        settings=[
            ["form_id", "form_title", "default_language"],
            ["my_form", "My Form Title", "en"],
        ],
    )
    assert data.title == "My Form Title"
    assert data.name == "my_form"


# ------------------------------------------------------------------ expression translation


@pytest.mark.parametrize(
    ("xpath", "expected"),
    [
        ("${field} = 1", "field == 1"),
        (". > 0", "value > 0"),
        ("${a} div ${b}", "a / b"),
        ("${a} mod 3", "a % 3"),
        ("${x} != ${y}", "x != y"),
    ],
)
def test_xpath_translation(xpath: str, expected: str) -> None:
    from app.importers.xlsform import _translate  # noqa: PLC0415

    assert _translate(xpath) == expected
