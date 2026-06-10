"""Unit tests for the tabular exporters and shared flatten logic (no DB required)."""

from __future__ import annotations

import csv
import io

from openpyxl import load_workbook

from app.exporters import compute_columns, export_csv, export_xlsx
from app.schemas.form_schema import FormSchema


def _form() -> FormSchema:
    """A form exercising scalar, multi_choice, and matrix fields across one page."""
    return FormSchema.model_validate(
        {
            "name": "survey",
            "title": "Survey",
            "pages": [
                {
                    "name": "p1",
                    "elements": [
                        {"type": "note", "name": "intro", "label": "Hi"},
                        {"type": "text", "name": "full_name", "label": "Name"},
                        {
                            "type": "multi_choice",
                            "name": "langs",
                            "label": "Languages",
                            "options": [
                                {"value": "py", "label": "Python"},
                                {"value": "go", "label": "Go"},
                            ],
                        },
                        {
                            "type": "matrix",
                            "name": "rate",
                            "label": "Rate",
                            "rows": [
                                {"value": "speed", "label": "Speed"},
                                {"value": "ease", "label": "Ease"},
                            ],
                            "columns": [
                                {"value": "low", "label": "Low"},
                                {"value": "high", "label": "High"},
                            ],
                        },
                    ],
                }
            ],
        }
    )


def _submissions() -> list[dict]:
    return [
        {
            "id": "s1",
            "created_at": "2026-06-08T10:00:00",
            "answers": {
                "full_name": "Ada",
                "langs": ["py", "go"],
                "rate": {"speed": "high", "ease": "low"},
            },
        },
        {
            "id": "s2",
            "created_at": "2026-06-08T11:00:00",
            "answers": {"full_name": "Linus", "langs": ["go"]},
        },
    ]


def test_compute_columns_layout() -> None:
    """Meta columns lead; notes are skipped; matrix expands one column per row."""
    cols = compute_columns(_form())
    assert cols == [
        "_id",
        "_submitted_at",
        "full_name",
        "langs",
        "rate/speed",
        "rate/ease",
    ]
    assert "intro" not in cols  # presentational note dropped


def test_csv_header_and_cells() -> None:
    text = export_csv(_form(), _submissions())
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)

    header = rows[0]
    assert header == ["_id", "_submitted_at", "full_name", "langs", "rate/speed", "rate/ease"]
    assert "_id" in header
    assert "rate/speed" in header and "rate/ease" in header

    by_id = {row[0]: row for row in rows[1:]}
    # multi_choice joined with "; "
    assert by_id["s1"][header.index("langs")] == "py; go"
    # matrix value lands in the matching per-row column
    assert by_id["s1"][header.index("rate/speed")] == "high"
    assert by_id["s1"][header.index("rate/ease")] == "low"
    # missing answers render as empty cells
    assert by_id["s2"][header.index("rate/speed")] == ""


def test_xlsx_reopens_with_expected_header() -> None:
    data = export_xlsx(_form(), _submissions())
    assert isinstance(data, bytes)

    workbook = load_workbook(io.BytesIO(data))
    sheet = workbook["Submissions"]
    header = [cell.value for cell in sheet[1]]
    assert header == ["_id", "_submitted_at", "full_name", "langs", "rate/speed", "rate/ease"]
    assert sheet[1][0].font.bold
    assert sheet.freeze_panes == "A2"

    second = {row[0].value: row for row in sheet.iter_rows(min_row=2)}
    langs_idx = header.index("langs")
    assert second["s1"][langs_idx].value == "py; go"


def _repeat_form() -> FormSchema:
    return FormSchema.model_validate(
        {
            "name": "household",
            "title": "Household",
            "pages": [
                {
                    "name": "p1",
                    "elements": [
                        {"type": "text", "name": "address", "label": "Address"},
                        {
                            "type": "repeat",
                            "name": "members",
                            "label": "Members",
                            "elements": [
                                {"type": "text", "name": "member_name", "label": "Name"},
                                {"type": "integer", "name": "age", "label": "Age"},
                            ],
                        },
                    ],
                }
            ],
        }
    )


def _repeat_submissions() -> list[dict]:
    return [
        {
            "id": "h1",
            "created_at": "2026-06-08T10:00:00",
            "answers": {
                "address": "1 Main St",
                "members": [
                    {"member_name": "Ada", "age": 40},
                    {"member_name": "Bo", "age": 12},
                ],
            },
        },
        {"id": "h2", "created_at": "2026-06-08T11:00:00", "answers": {"address": "2 Oak Ave"}},
    ]


def test_main_sheet_treats_repeat_as_a_single_column() -> None:
    # Repeat child fields must NOT leak in as (always-empty) top-level columns.
    cols = compute_columns(_repeat_form())
    assert cols == ["_id", "_submitted_at", "address", "members"]
    assert "member_name" not in cols and "age" not in cols


def test_xlsx_emits_one_long_format_sheet_per_repeat() -> None:
    data = export_xlsx(_repeat_form(), _repeat_submissions())
    workbook = load_workbook(io.BytesIO(data))
    assert "Submissions" in workbook.sheetnames
    assert "members" in workbook.sheetnames

    members = workbook["members"]
    header = [cell.value for cell in members[1]]
    assert header == ["_parent_id", "_index", "member_name", "age"]

    rows = [[c.value for c in row] for row in members.iter_rows(min_row=2)]
    # Two instances from h1, linked back to the parent submission, in order; none from h2.
    assert rows == [
        ["h1", "0", "Ada", "40"],
        ["h1", "1", "Bo", "12"],
    ]
