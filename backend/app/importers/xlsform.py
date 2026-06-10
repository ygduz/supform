"""XLSForm  ->  Supform schema importer.

XLSForm (https://xlsform.org) is the spreadsheet standard KoboToolbox and ODK use. A
workbook has ``survey``, ``choices`` and ``settings`` sheets. This importer maps that grid
onto Supform's richer, nested JSON model so forms can migrate in from the ODK ecosystem.

Mapping (see docs/form-schema.md for the full table):

    XLSForm type            -> Supform element type
    -----------------------------------------------------
    text                    -> text
    integer / decimal       -> integer / decimal
    select_one <list>       -> single_choice  (options from `choices` list)
    select_multiple <list>  -> multi_choice
    note                    -> note
    begin group / end       -> group (nested elements)
    begin repeat / end      -> repeat (nested elements)
    geopoint                -> geopoint
    image / audio / video   -> image / file / file
    calculate               -> calculated
    range                   -> scale

Column mapping: ``relevant`` -> ``visibleIf``; ``constraint`` (+ ``constraint_message``)
-> ``validation``; ``calculation`` -> ``calculate``; ``required`` -> ``required``.

ODK XPath expressions are translated to Supform's expression grammar on a best-effort
basis (``${field}`` -> ``field``, ``.`` -> ``value``, ``=`` -> ``==``, ``div``/``mod`` ->
``/``/``%``, ``selected()`` is kept). Complex XPath may need manual review after import.
"""

from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import load_workbook

from app.schemas.form_schema import FormSchema

Row = dict[str, Any]

# XLSForm base type -> Supform element type. select_one/select_multiple handled separately.
TYPE_MAP: dict[str, str] = {
    "text": "text",
    "string": "text",
    "integer": "integer",
    "int": "integer",
    "decimal": "decimal",
    "note": "note",
    "date": "date",
    "time": "time",
    "datetime": "datetime",
    "geopoint": "geopoint",
    "image": "image",
    "photo": "image",
    "audio": "file",
    "video": "file",
    "file": "file",
    "barcode": "barcode",
    "calculate": "calculated",
    "range": "scale",
    "acknowledge": "boolean",
    "email": "email",
    "url": "url",
}

_BEGIN_GROUP = {"begin group", "begin_group"}
_END_GROUP = {"end group", "end_group"}
_BEGIN_REPEAT = {"begin repeat", "begin_repeat"}
_END_REPEAT = {"end repeat", "end_repeat"}
_TRUTHY = {"yes", "true", "1", "true()"}


def import_xlsform(source: str | Path | bytes | BinaryIO) -> FormSchema:
    """Parse an ``.xlsx`` XLSForm into a :class:`FormSchema`."""
    stream: Any = io.BytesIO(source) if isinstance(source, bytes) else source
    wb = load_workbook(stream, read_only=True, data_only=True)

    settings = _read_settings(wb)
    default_lang = settings.get("default_language", "en")
    choices = _read_choices(wb, default_lang)
    survey_rows, label_langs = _read_sheet(wb, "survey")

    elements = _build_elements(survey_rows, choices, default_lang)

    form: dict[str, Any] = {
        "schemaVersion": "1.0",
        "name": _slug(settings.get("form_id") or settings.get("form_title") or "imported_form"),
        "title": settings.get("form_title") or "Imported form",
        "defaultLanguage": default_lang,
        "pages": [{"name": "page1", "elements": elements}],
    }
    if label_langs:
        form["languages"] = sorted(label_langs)
    return FormSchema.model_validate(form)


# ---------------------------------------------------------------- sheet reading
def _read_sheet(wb: Any, name: str) -> tuple[list[Row], set[str]]:
    """Return (rows-as-dicts, set-of-label-languages). Missing sheet -> ([], set())."""
    if name not in wb.sheetnames:
        return [], set()
    ws = wb[name]
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return [], set()
    cols = [str(c).strip() if c is not None else "" for c in header]
    langs = {lang for c in cols if (lang := _label_lang(c))}
    rows: list[Row] = []
    for raw in it:
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        rows.append({cols[i]: raw[i] for i in range(min(len(cols), len(raw))) if cols[i]})
    return rows, langs


def _read_settings(wb: Any) -> dict[str, str]:
    rows, _ = _read_sheet(wb, "settings")
    if not rows:
        return {}
    return {k: str(v).strip() for k, v in rows[0].items() if v is not None}


def _read_choices(wb: Any, default_lang: str) -> dict[str, list[dict[str, Any]]]:
    rows, _ = _read_sheet(wb, "choices")
    lists: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        list_name = _cell(row, "list_name") or _cell(row, "list name")
        value = _cell(row, "name")
        if not list_name or value is None:
            continue
        choice: dict[str, Any] = {"value": value}
        label = _label(row, default_lang)
        if label is not None:
            choice["label"] = label
        lists.setdefault(str(list_name), []).append(choice)
    return lists


# ------------------------------------------------------------- survey -> tree
def _build_elements(
    rows: list[Row], choices: dict[str, list[dict[str, Any]]], default_lang: str
) -> list[dict[str, Any]]:
    root: list[dict[str, Any]] = []
    stack: list[list[dict[str, Any]]] = [root]  # children list of the current container

    for row in rows:
        type_raw = str(_cell(row, "type") or "").strip()
        if not type_raw:
            continue
        low = type_raw.lower()

        if low in _END_GROUP or low in _END_REPEAT:
            if len(stack) > 1:
                stack.pop()
            continue

        if low in _BEGIN_GROUP:
            el = _container(row, "group", default_lang)
            stack[-1].append(el)
            stack.append(el["elements"])
            continue

        if low in _BEGIN_REPEAT:
            el = _container(row, "repeat", default_lang)
            el["repeat"] = {"min": 0}
            stack[-1].append(el)
            stack.append(el["elements"])
            continue

        field = _field(type_raw, row, choices, default_lang)
        if field is not None:
            stack[-1].append(field)

    return root


def _container(row: Row, kind: str, default_lang: str) -> dict[str, Any]:
    el: dict[str, Any] = {"type": kind, "name": _slug(_cell(row, "name") or kind), "elements": []}
    label = _label(row, default_lang)
    if label is not None:
        el["label"] = label
    _apply_logic(el, row)
    return el


def _field(
    type_raw: str,
    row: Row,
    choices: dict[str, list[dict[str, Any]]],
    default_lang: str,
) -> dict[str, Any] | None:
    parts = type_raw.split()
    base = parts[0].lower()
    # Support both "select_one list" and "select one list".
    if base in ("select_one", "select") and len(parts) >= 2 and parts[1].lower() == "one":
        base, list_name = "select_one", parts[2] if len(parts) > 2 else ""
    elif (
        base in ("select_multiple", "select")
        and len(parts) >= 2
        and parts[1].lower()
        in (
            "multiple",
            "multi",
        )
    ):
        base, list_name = "select_multiple", parts[2] if len(parts) > 2 else ""
    else:
        list_name = parts[1] if len(parts) > 1 else ""

    name = _slug(_cell(row, "name") or "field")
    el: dict[str, Any] = {"name": name}
    label = _label(row, default_lang)
    if label is not None:
        el["label"] = label
    hint = _label(row, default_lang, prefix="hint")
    if hint is not None:
        el["hint"] = hint

    if base == "select_one":
        el["type"] = "single_choice"
        el["options"] = choices.get(list_name, [])
    elif base == "select_multiple":
        el["type"] = "multi_choice"
        el["options"] = choices.get(list_name, [])
    else:
        mapped = TYPE_MAP.get(base)
        if mapped is None:
            return None  # unsupported types (e.g. 'start', 'end' metadata) are dropped
        el["type"] = mapped

    calc = _cell(row, "calculation")
    if el["type"] == "calculated" or calc:
        if calc:
            el["calculate"] = _translate(str(calc))
        el.setdefault("type", "calculated")
        el["readOnly"] = True

    _apply_logic(el, row)
    return el


def _apply_logic(el: dict[str, Any], row: Row) -> None:
    if _truthy(_cell(row, "required")):
        el["required"] = True
    relevant = _cell(row, "relevant")
    if relevant:
        el["visibleIf"] = _translate(str(relevant))
    constraint = _cell(row, "constraint")
    if constraint:
        validation: dict[str, Any] = {"expression": _translate(str(constraint))}
        message = _cell(row, "constraint_message") or _cell(row, "constraint message")
        if message:
            validation["message"] = str(message)
        el["validation"] = validation


# ----------------------------------------------------------------- helpers
def _cell(row: Row, key: str) -> Any:
    val = row.get(key)
    return val if val != "" else None


def _label_lang(column: str) -> str | None:
    """Return the language code embedded in a 'label::English (en)' style column, if any."""
    if "::" not in column:
        return None
    suffix = column.split("::", 1)[1].strip()
    match = re.search(r"\(([^)]+)\)\s*$", suffix)
    return (match.group(1) if match else suffix).strip() or None


def _label(row: Row, default_lang: str, *, prefix: str = "label") -> Any:
    """Resolve a label/hint cell to a string or an {lang: text} i18n map."""
    translations: dict[str, str] = {}
    plain: str | None = None
    for col, val in row.items():
        if val is None or str(val).strip() == "":
            continue
        if col == prefix:
            plain = str(val)
        elif col.startswith(f"{prefix}::"):
            lang = _label_lang(col) or default_lang
            translations[lang] = str(val)
    if translations:
        if plain is not None:
            translations.setdefault(default_lang, plain)
        return translations
    return plain


def _truthy(value: Any) -> bool:
    return value is not None and str(value).strip().lower() in _TRUTHY


def _slug(value: Any) -> str:
    s = re.sub(r"[^a-zA-Z0-9_]", "_", str(value).strip())
    if not s:
        s = "field"
    if not re.match(r"[a-zA-Z_]", s[0]):
        s = f"_{s}"
    return s


def _translate(expr: str) -> str:
    """Best-effort translation of an ODK/XPath expression to Supform's grammar."""
    s = expr.strip()
    s = re.sub(r"\$\{([^}]+)\}", r"\1", s)  # ${field} -> field
    s = re.sub(r"(?<![\w.])\.(?![\w.])", "value", s)  # current node '.' -> value
    s = re.sub(r"\bdiv\b", "/", s)
    s = re.sub(r"\bmod\b", "%", s)
    s = re.sub(r"\bnot\s*\(", "not (", s)
    s = re.sub(r"(?<![<>=!])=(?!=)", "==", s)  # '=' -> '==' (leave <=,>=,!=,==)
    return s
