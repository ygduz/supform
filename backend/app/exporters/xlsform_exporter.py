"""XLSForm export — turn a Supform schema into an ODK-compatible XLSForm workbook.

Produces the standard three-sheet layout (survey / choices / settings) that ODK Collect,
KoboToolbox, and SurveyCTO can open directly.  This is the structural inverse of the
XLSForm importer (app/importers/xlsform.py).

Unsupported Supform concepts and their best-effort fallbacks
------------------------------------------------------------
``matrix``      → one ``select_one`` per row, named ``{matrix_name}_{row_value}``.
``rating``      → ``range`` (min 1, max value of last option, or 5).
``scale``       → ``range``.
``longtext``    → ``text`` (no separate XLSForm type).
``phone``       → ``text`` (XLSForm has no phone type).
``signature``   → ``image``.
``dropdown``    → ``select_one`` (same semantics).
``ranking``     → ``rank`` (ODK has a rank type since XLSForm 2.x).
``html/note``   → ``note`` (HTML stripped out, plain text shown).
``section``     → skipped (presentational only in Supform).
``geotrace``    → ``geotrace``.
``geoshape``    → ``geoshape``.
``hidden``      → ``hidden``.
``calculated``  → ``calculate``.
"""

from __future__ import annotations

import io
import re
from collections.abc import Iterable
from typing import Any

from app.schemas.form_schema import Element, ElementType, FormSchema

# --------------------------------------------------------------------------- #
# Type mapping: Supform → XLSForm survey type string
# --------------------------------------------------------------------------- #

_SIMPLE_TYPE: dict[str, str] = {
    ElementType.TEXT: "text",
    ElementType.LONGTEXT: "text",
    ElementType.EMAIL: "email",
    ElementType.URL: "text",
    ElementType.PHONE: "text",
    ElementType.NUMBER: "decimal",
    ElementType.INTEGER: "integer",
    ElementType.DECIMAL: "decimal",
    ElementType.DATE: "date",
    ElementType.TIME: "time",
    ElementType.DATETIME: "datetime",
    ElementType.BOOLEAN: "acknowledge",
    ElementType.GEOPOINT: "geopoint",
    ElementType.GEOTRACE: "geotrace",
    ElementType.GEOSHAPE: "geoshape",
    ElementType.FILE: "file",
    ElementType.IMAGE: "image",
    ElementType.SIGNATURE: "image",
    ElementType.BARCODE: "barcode",
    ElementType.NOTE: "note",
    ElementType.CALCULATED: "calculate",
    ElementType.HIDDEN: "hidden",
    ElementType.START: "start",
    ElementType.END: "end",
    ElementType.TODAY: "today",
    ElementType.DEVICEID: "deviceid",
    ElementType.USERNAME: "username",
}

_PRESENTATIONAL = {ElementType.SECTION, ElementType.HTML}


# --------------------------------------------------------------------------- #
# Expression back-translation (mirror of importer's _translate)
# --------------------------------------------------------------------------- #


def _expr_to_xlsform(expr: str | None) -> str:
    """Convert a Supform expression back to XLSForm/XPath syntax."""
    if not expr:
        return ""
    # Wrap field references in ${...}  — but skip True/False/None keywords.
    result = re.sub(
        r"\b([A-Za-z_]\w*)\b",
        lambda m: (
            m.group(0)
            if m.group(0)
            in {"and", "or", "not", "True", "False", "None", "selected", "count", "min", "max"}
            else f"${{{m.group(0)}}}"
        ),
        expr,
    )
    result = result.replace("==", "=").replace("!=", "!=")  # != stays, == → =
    # Logical operators
    result = result.replace(" and ", " and ").replace(" or ", " or ")
    return result


# --------------------------------------------------------------------------- #
# Label helpers
# --------------------------------------------------------------------------- #


def _localize(value: Any, lang: str = "default") -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, dict):
        return value.get(lang) or next(iter(value.values()), "") or ""
    return str(value)


def _all_languages(schema: FormSchema) -> list[str]:
    langs = list(schema.languages or [])
    default = schema.default_language or "en"
    if default not in langs:
        langs.insert(0, default)
    return langs


# --------------------------------------------------------------------------- #
# Survey rows builder
# --------------------------------------------------------------------------- #

ChoiceRow = dict[str, str]
SurveyRow = dict[str, str]


def _survey_rows(
    elements: list[Element],
    choice_lists: dict[str, list[ChoiceRow]],
    langs: list[str],
) -> list[SurveyRow]:
    rows: list[SurveyRow] = []

    for el in elements:
        if el.type in _PRESENTATIONAL:
            continue

        # --- Containers ---
        if el.type == ElementType.GROUP:
            row = _base_row(el, langs)
            row["type"] = "begin_group"
            rows.append(row)
            rows.extend(_survey_rows(el.elements or [], choice_lists, langs))
            rows.append({"type": "end_group", "name": el.name})
            continue

        if el.type == ElementType.REPEAT:
            row = _base_row(el, langs)
            row["type"] = "begin_repeat"
            if el.repeat and el.repeat.max is not None:
                row["repeat_count"] = str(el.repeat.max)
            rows.append(row)
            rows.extend(_survey_rows(el.elements or [], choice_lists, langs))
            rows.append({"type": "end_repeat", "name": el.name})
            continue

        # --- Matrix: expand to one select_one per row ---
        if el.type == ElementType.MATRIX:
            list_name = f"_choices_{el.name}_cols"
            if el.columns and list_name not in choice_lists:
                choice_lists[list_name] = [
                    {"list_name": list_name, "name": str(c.value), **_label_cols(c.label, langs)}
                    for c in el.columns
                ]
            for mrow in el.rows or []:
                sub_name = f"{el.name}_{mrow.value}"
                sub_label = f"{_localize(el.label)} — {_localize(mrow.label)}"
                row: SurveyRow = {
                    "type": f"select_one {list_name}",
                    "name": sub_name,
                    "label": sub_label,
                }
                if el.required:
                    row["required"] = "yes"
                rows.append(row)
            continue

        # --- Scale / Rating → range ---
        if el.type in {ElementType.SCALE, ElementType.RATING}:
            row = _base_row(el, langs)
            opts = el.options or []
            lo = opts[0].value if opts else 1
            hi = opts[-1].value if opts else 5
            row["type"] = "range"
            row["parameters"] = f"start={lo} end={hi} step=1"
            rows.append(row)
            continue

        # --- Choice fields ---
        if el.type in {ElementType.SINGLE_CHOICE, ElementType.DROPDOWN}:
            list_name = _register_choices(el, choice_lists, langs)
            row = _base_row(el, langs)
            row["type"] = f"select_one {list_name}"
            rows.append(row)
            continue

        if el.type == ElementType.MULTI_CHOICE:
            list_name = _register_choices(el, choice_lists, langs)
            row = _base_row(el, langs)
            row["type"] = f"select_multiple {list_name}"
            rows.append(row)
            continue

        if el.type == ElementType.RANKING:
            list_name = _register_choices(el, choice_lists, langs)
            row = _base_row(el, langs)
            row["type"] = f"rank {list_name}"
            rows.append(row)
            continue

        # --- Simple types ---
        xlsform_type = _SIMPLE_TYPE.get(el.type, "text")
        row = _base_row(el, langs)
        row["type"] = xlsform_type
        rows.append(row)

    return rows


def _base_row(el: Element, langs: list[str]) -> SurveyRow:
    row: SurveyRow = {"type": "", "name": el.name}

    # Label — single or multilingual
    row.update(_label_cols(el.label, langs))
    if el.hint:
        row.update(_hint_cols(el.hint, langs))
    if el.required:
        row["required"] = "yes"
    if el.read_only:
        row["read_only"] = "yes"
    if el.default_value is not None:
        row["default"] = str(el.default_value)
    if el.visible_if:
        row["relevant"] = _expr_to_xlsform(el.visible_if)
    if el.calculate:
        row["calculation"] = _expr_to_xlsform(el.calculate)
    if el.validation:
        v = el.validation
        parts: list[str] = []
        if v.expression:
            parts.append(_expr_to_xlsform(v.expression))
        if v.min is not None:
            parts.append(f". >= {v.min}")
        if v.max is not None:
            parts.append(f". <= {v.max}")
        if v.min_length is not None:
            parts.append(f"string-length(.) >= {v.min_length}")
        if v.max_length is not None:
            parts.append(f"string-length(.) <= {v.max_length}")
        if parts:
            row["constraint"] = " and ".join(parts)
        if v.message:
            row["constraint_message"] = _localize(v.message)

    return row


def _label_cols(label: Any, langs: list[str]) -> dict[str, str]:
    if label is None:
        return {"label": ""}
    if isinstance(label, str):
        return {"label": label}
    # Multilingual: emit label::Language (code) for each language
    result: dict[str, str] = {}
    for lang in langs:
        col = f"label::{lang}"
        result[col] = label.get(lang, "")
    if not result:
        result["label"] = ""
    return result


def _hint_cols(hint: Any, langs: list[str]) -> dict[str, str]:
    if hint is None:
        return {}
    if isinstance(hint, str):
        return {"hint": hint}
    result: dict[str, str] = {}
    for lang in langs:
        result[f"hint::{lang}"] = hint.get(lang, "")
    return result


def _register_choices(
    el: Element,
    choice_lists: dict[str, list[ChoiceRow]],
    langs: list[str],
) -> str:
    list_name = f"_choices_{el.name}"
    if list_name not in choice_lists and el.options:
        choice_lists[list_name] = [
            {"list_name": list_name, "name": str(c.value), **_label_cols(c.label, langs)}
            for c in el.options
        ]
    return list_name


# --------------------------------------------------------------------------- #
# Public API
# --------------------------------------------------------------------------- #


def export_xlsform(form: FormSchema, _submissions: Iterable[dict[str, Any]] = ()) -> bytes:
    """Return an XLSForm workbook as bytes (XLSX).

    The submissions argument is accepted for API-consistency but ignored —
    XLSForm describes the form structure, not collected data.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for XLSForm export") from exc

    langs = _all_languages(form)
    choice_lists: dict[str, list[ChoiceRow]] = {}

    all_elements: list[Element] = []
    for page in form.pages:
        all_elements.extend(page.elements)

    survey_data = _survey_rows(all_elements, choice_lists, langs)

    # --- Derive column sets ---
    survey_cols: list[str] = ["type", "name"]
    # Collect all label/hint/etc columns that appear in any row
    extra: list[str] = []
    for row in survey_data:
        for k in row:
            if k not in survey_cols and k not in extra:
                extra.append(k)
    survey_cols.extend(extra)

    # Standard ODK ordering — put common columns first
    pref = [
        "label",
        "hint",
        "required",
        "relevant",
        "constraint",
        "constraint_message",
        "calculation",
        "read_only",
        "default",
        "appearance",
        "parameters",
        "repeat_count",
    ]
    ordered_extra = [c for c in pref if c in extra or any(c in r for r in survey_data)]
    remaining = [c for c in extra if c not in ordered_extra]
    # Also include language-specific label columns after base label
    lang_label_cols = [c for c in remaining if c.startswith("label::") or c.startswith("hint::")]
    other_cols = [c for c in remaining if c not in lang_label_cols]
    survey_cols = ["type", "name"] + ordered_extra + lang_label_cols + other_cols

    choices_cols: list[str] = ["list_name", "name"]
    for rows in choice_lists.values():
        for row in rows:
            for k in row:
                if k not in choices_cols:
                    choices_cols.append(k)

    # --- Build workbook ---
    wb = openpyxl.Workbook()
    # Remove the default empty sheet openpyxl always creates.
    del wb["Sheet"]
    _write_sheet(wb, "survey", survey_cols, survey_data)
    _write_sheet(wb, "choices", choices_cols, [r for rows in choice_lists.values() for r in rows])
    _write_settings(wb, form, langs)

    # --- Style header rows ---
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        sheet.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_sheet(wb: Any, name: str, columns: list[str], rows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet(name)
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])


def _write_settings(wb: Any, form: FormSchema, langs: list[str]) -> None:
    ws = wb.create_sheet("settings")
    cols = ["form_title", "form_id", "version", "default_language"]
    ws.append(cols)
    ws.append(
        [
            _localize(form.title),
            form.name,
            str(form.version),
            form.default_language or (langs[0] if langs else "en"),
        ]
    )
